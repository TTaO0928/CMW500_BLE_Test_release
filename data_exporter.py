"""
CMW500 自动化测试工具 - 数据导出模块

功能说明：
    将 BLE TX 调制 / TX 功率 / RX PER 测试结果导出为 Excel 文件。
    使用 pandas 处理数据，openpyxl 设置样式。
    文件名自动包含日期时间戳，避免覆盖历史数据。

输出内容：
    - "TX 调制数据" / "TX 调制摘要"
    - "TX 功率数据" / "TX 功率摘要"
    - "RX PER 数据" / "RX PER 测试摘要"
"""

import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from test_executor import MODULATION_ITEMS, POWER_ITEMS


class DataExporter:
    """测试结果数据导出类"""

    # 预定义样式常量
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 浅绿色
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 浅红色
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # 蓝色
    HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)                 # 白色加粗
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    BOLD_FONT = Font(name="微软雅黑", bold=True, size=10)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, config):
        """
        初始化数据导出器

        参数:
            config: 从 config.yaml 加载的配置字典
        """
        self.export_config = config["export"]
        self.file_prefix = self.export_config["file_prefix"]

        # 解析输出目录路径：相对路径基于程序所在目录（兼容 exe 打包）
        raw_output_dir = self.export_config["output_dir"]
        if os.path.isabs(raw_output_dir):
            self.output_dir = raw_output_dir
        else:
            # 获取程序根目录（兼容 PyInstaller 打包）
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            self.output_dir = os.path.join(app_dir, raw_output_dir)

    def _ensure_output_dir(self):
        """确保输出目录存在，不存在则创建"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def _generate_filename(self):
        """
        生成带日期时间戳的文件名

        格式示例：BLE_TX_Modulation_Test_20260702_164643.xlsx

        返回:
            str: 完整的文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.file_prefix}_{timestamp}.xlsx"
        return os.path.join(self.output_dir, filename)

    def export_to_excel(self, results, test_params):
        """
        将测试结果导出为格式化的 Excel 文件

        参数:
            results:     测试结果列表
            test_params: 测试参数配置字典

        返回:
            str: 导出文件的完整路径
        """
        self._ensure_output_dir()

        tx_mod_results = [r for r in results if r.get("test_type") == "tx_modulation"]
        tx_pow_results = [r for r in results if r.get("test_type") == "tx_power"]
        rx_results = [r for r in results if r.get("test_type") == "rx_per"]

        # 动态选择文件名前缀
        if tx_mod_results and tx_pow_results and rx_results:
            prefix = "CMW500_BLE_All_Test"
        elif tx_mod_results and tx_pow_results:
            prefix = "CMW500_BLE_TX_Test"
        elif rx_results:
            prefix = "CMW500_BLE_RX_PER_Test"
        else:
            prefix = self.file_prefix

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            if tx_mod_results:
                self._write_tx_modulation_sheet(writer, tx_mod_results, test_params)
            if tx_pow_results:
                self._write_tx_power_sheet(writer, tx_pow_results, test_params)
            if rx_results:
                self._write_rx_per_sheet(writer, rx_results, test_params)

        # 样式应用
        self._apply_styles(filepath)
        return filepath

    def _build_tx_modulation_dataframe(self, results, test_params):
        """构建 TX 调制测试结果 DataFrame"""
        rows = []
        mod_cfg = test_params.get("modulation_measurements", {})
        for r in results:
            row = {
                "信道 (Channel)": r.get("channel", ""),
                "测量时间": r.get("timestamp", ""),
            }
            for key, info in MODULATION_ITEMS.items():
                if key not in mod_cfg:
                    continue
                name = mod_cfg[key].get("name", info[3])
                unit = mod_cfg[key].get("unit", info[4])
                row[f"{name} ({unit})"] = r.get(key, "N/A")
                if "pass_fail" in r:
                    row[f"{name} 判定"] = r["pass_fail"].get(key, "N/A")
                else:
                    row[f"{name} 判定"] = "ERROR"
            rows.append(row)
        return pd.DataFrame(rows)

    def _build_tx_power_dataframe(self, results, test_params):
        """构建 TX 功率测试结果 DataFrame"""
        rows = []
        pow_cfg = test_params.get("power_measurements", {})
        for r in results:
            row = {
                "信道 (Channel)": r.get("channel", ""),
                "测量时间": r.get("timestamp", ""),
            }
            for key, info in POWER_ITEMS.items():
                if key not in pow_cfg:
                    continue
                name = pow_cfg[key].get("name", info[2])
                unit = pow_cfg[key].get("unit", info[3])
                row[f"{name} ({unit})"] = r.get(key, "N/A")
                if "pass_fail" in r:
                    row[f"{name} 判定"] = r["pass_fail"].get(key, "N/A")
                else:
                    row[f"{name} 判定"] = "ERROR"
            rows.append(row)
        return pd.DataFrame(rows)

    def _write_tx_modulation_sheet(self, writer, results, test_params):
        """写入 TX 调制测试结果到 Excel"""
        df = self._build_tx_modulation_dataframe(results, test_params)
        df.to_excel(writer, sheet_name="TX 调制数据", index=False)
        summary_data = self._build_modulation_summary(results, test_params)
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="TX 调制摘要", index=False)

    def _write_tx_power_sheet(self, writer, results, test_params):
        """写入 TX 功率测试结果到 Excel"""
        df = self._build_tx_power_dataframe(results, test_params)
        df.to_excel(writer, sheet_name="TX 功率数据", index=False)
        summary_data = self._build_power_summary(results, test_params)
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="TX 功率摘要", index=False)

    def _build_rx_per_dataframe(self, results, test_params):
        """构建 RX PER 测试结果 DataFrame"""
        rx_cfg = test_params.get("rx_per", {})
        rows = []
        for r in results:
            rows.append({
                "信道 (Channel)": r.get("channel", ""),
                "测量时间": r.get("timestamp", ""),
                "最佳灵敏度 (dBm)": r.get("sensitivity", "N/A"),
                "PER 阈值 (%)": rx_cfg.get("per_threshold", 30.8),
            })
        return pd.DataFrame(rows)

    def _write_rx_per_sheet(self, writer, results, test_params):
        """写入 RX PER 测试结果到 Excel"""
        df = self._build_rx_per_dataframe(results, test_params)
        df.to_excel(writer, sheet_name="RX PER 数据", index=False)
        summary_data = self._build_rx_per_summary(results, test_params)
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="RX PER 测试摘要", index=False)

    def _build_common_summary_header(self, test_params):
        """构建通用摘要头部"""
        return [
            {"项目": "测试时间", "数值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"项目": "测试标准", "数值": f"{test_params.get('standard', 'BT_LE')} ({test_params.get('phy_type', 'LE_1Msps')})"},
            {"项目": "统计次数", "数值": test_params.get("statistic_count", 10)},
        ]

    def _build_modulation_summary(self, results, test_params):
        """构建 TX 调制测试摘要数据"""
        total_channels = len(results)
        mod_cfg = test_params.get("modulation_measurements", {})

        summary_rows = self._build_common_summary_header(test_params)
        summary_rows.append({"项目": "总测试信道数", "数值": total_channels})
        summary_rows.append({"项目": "———", "数值": "———"})

        for key, info in MODULATION_ITEMS.items():
            if key not in mod_cfg:
                continue
            cfg = mod_cfg[key]
            name = cfg.get("name", info[3])
            unit = cfg.get("unit", info[4])
            upper = cfg.get("upper_limit")
            lower = cfg.get("lower_limit")

            pass_count = 0
            fail_count = 0
            for r in results:
                if "pass_fail" in r:
                    if r["pass_fail"].get(key) == "PASS":
                        pass_count += 1
                    else:
                        fail_count += 1

            limit_text = f"上限: {upper} {unit}" if upper is not None else ""
            if lower is not None:
                limit_text += f" 下限: {lower} {unit}" if limit_text else f"下限: {lower} {unit}"

            summary_rows.append({
                "项目": f"{name} ({limit_text})",
                "数值": f"通过 {pass_count} / 失败 {fail_count}",
            })

        all_pass = sum(
            1 for r in results
            if "pass_fail" in r and all(v == "PASS" for v in r["pass_fail"].values())
        )
        summary_rows.append({"项目": "———", "数值": "———"})
        summary_rows.append({"项目": "全部通过信道数", "数值": all_pass})
        summary_rows.append({"项目": "有失败项信道数", "数值": total_channels - all_pass})
        summary_rows.append({"项目": "总体判定", "数值": "PASS" if all_pass == total_channels else "FAIL"})

        return summary_rows

    def _build_power_summary(self, results, test_params):
        """构建 TX 功率测试摘要数据"""
        total_channels = len(results)
        pow_cfg = test_params.get("power_measurements", {})

        summary_rows = self._build_common_summary_header(test_params)
        summary_rows.append({"项目": "总测试信道数", "数值": total_channels})
        summary_rows.append({"项目": "———", "数值": "———"})

        for key, info in POWER_ITEMS.items():
            if key not in pow_cfg:
                continue
            cfg = pow_cfg[key]
            name = cfg.get("name", info[2])
            unit = cfg.get("unit", info[3])
            upper = cfg.get("upper_limit")
            lower = cfg.get("lower_limit")

            pass_count = 0
            fail_count = 0
            for r in results:
                if "pass_fail" in r:
                    if r["pass_fail"].get(key) == "PASS":
                        pass_count += 1
                    else:
                        fail_count += 1

            limit_text = f"上限: {upper} {unit}" if upper is not None else ""
            if lower is not None:
                limit_text += f" 下限: {lower} {unit}" if limit_text else f"下限: {lower} {unit}"

            summary_rows.append({
                "项目": f"{name} ({limit_text or '仅记录'})",
                "数值": f"通过 {pass_count} / 失败 {fail_count}",
            })

        all_pass = sum(
            1 for r in results
            if "pass_fail" in r and all(v == "PASS" for v in r["pass_fail"].values())
        )
        summary_rows.append({"项目": "———", "数值": "———"})
        summary_rows.append({"项目": "全部通过信道数", "数值": all_pass})
        summary_rows.append({"项目": "有失败项信道数", "数值": total_channels - all_pass})
        summary_rows.append({"项目": "总体判定", "数值": "PASS" if all_pass == total_channels else "FAIL"})

        return summary_rows

    def _build_rx_per_summary(self, results, test_params):
        """
        构建 RX PER 测试摘要数据（无 PASS/FAIL 判定）

        参数:
            results:     RX PER 测试结果列表
            test_params: 测试参数配置

        返回:
            list[dict]: 摘要行列表
        """
        rx_cfg = test_params.get("rx_per", {})
        total_channels = len(results)
        found_count = sum(1 for r in results if r.get("sensitivity") is not None)

        summary_rows = [
            {"项目": "测试时间", "数值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"项目": "测试类型", "数值": "BLE RX PER 接收灵敏度搜索"},
            {"项目": "信道范围", "数值": f"Ch {test_params['channel_start']} ~ Ch {test_params['channel_end']}"},
            {"项目": "起始功率 (dBm)", "数值": rx_cfg.get("start_power", -90.0)},
            {"项目": "结束功率 (dBm)", "数值": rx_cfg.get("end_power", -100.0)},
            {"项目": "功率步进 (dBm)", "数值": rx_cfg.get("step_size", 0.5)},
            {"项目": "Expected Nominal Power (dBm)", "数值": rx_cfg.get("exp_nom_pow", 20.0)},
            {"项目": "PER 阈值 (%)", "数值": rx_cfg.get("per_threshold", 30.8)},
            {"项目": "每点发包数", "数值": rx_cfg.get("packet_count", 500)},
            {"项目": "———", "数值": "———"},
            {"项目": "总测试信道数", "数值": total_channels},
            {"项目": "找到灵敏度点数", "数值": found_count},
        ]
        return summary_rows

    def _apply_styles(self, filepath):
        """
        对导出的 Excel 文件应用样式美化

        参数:
            filepath: Excel 文件路径
        """
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            is_summary = "摘要" in sheet_name

            # 设置表头样式
            for cell in ws[1]:
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.CENTER_ALIGN
                cell.border = self.THIN_BORDER

            # 设置数据区域样式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = self.NORMAL_FONT
                    cell.border = self.THIN_BORDER
                    cell.alignment = self.CENTER_ALIGN

                    # 对判定值着色
                    if cell.value == "PASS":
                        cell.fill = self.PASS_FILL
                    elif cell.value == "FAIL":
                        cell.fill = self.FAIL_FILL

            # 自动调整列宽
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_length = max(max_length, cell_length)
                ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

            # 摘要 Sheet 加宽关键列
            if is_summary:
                ws.column_dimensions["A"].width = 35
                ws.column_dimensions["B"].width = 30

        # 保存样式
        wb.save(filepath)
